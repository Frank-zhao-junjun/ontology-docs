from flask import Blueprint, jsonify, Response, request
from .auth import require_auth
from .models import MetaModelDefinition, MetaModelRelease, MetaModelReleaseItem, Domain
import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

bp_export = Blueprint('export', __name__, url_prefix='/api/export')
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")


def _get_release_models(release_no: str):
    rel = MetaModelRelease.query.filter_by(release_no=release_no).first()
    if not rel:
        return None, None
    items = MetaModelReleaseItem.query.filter_by(release_id=rel.id).all()
    ids = [i.model_definition_id for i in items]
    models = MetaModelDefinition.query.filter(MetaModelDefinition.id.in_(ids)).all()
    result = {
        'version': release_no,
        'exported_at': rel.released_at.isoformat() if rel.released_at else None,
    }
    for m in models:
        result[m.model_type] = m.content_json
    return result, rel


@bp_export.get('/json/<release_no>')
@require_auth()
def export_json(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    return jsonify(result)


@bp_export.get('/yaml/<release_no>')
@require_auth()
def export_yaml(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    yaml_str = yaml.dump(result, allow_unicode=True, default_flow_style=False, sort_keys=False)
    return Response(yaml_str, mimetype='text/yaml')


def _write_sheet(ws, headers: list[str], rows: list[list[str]]):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val) if val is not None else '')


def _build_xlsx(result: dict) -> BytesIO:
    """从 result dict 生成多 Sheet Excel 工作簿。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: 领域总览
    ws1 = wb.create_sheet('领域总览')
    domains = Domain.query.all()
    _write_sheet(ws1, ['名称', '描述', '标签'], [[d.name, d.description or '', d.tags or ''] for d in domains])

    # Sheet 2: 维1_静态结构
    s = result.get('structural', {})
    entities = s.get('entities', [])
    rows_v1 = []
    for e in entities:
        for a in e.get('attributes', []):
            rows_v1.append([e.get('id',''), e.get('name',''), a.get('id',''), a.get('type',''), str(a.get('required','')), str(a.get('unique',''))])
    for r in s.get('relations', []):
        rows_v1.append(['', '', r.get('type',''), r.get('source',''), r.get('target',''), r.get('inverseOf','')])
    ws2 = wb.create_sheet('维1_静态结构')
    _write_sheet(ws2, ['实体ID', '实体名', '属性/关系', '类型', '必填', '唯一/inverseOf'], rows_v1)

    # Sheet 3: 维2_动态行为
    b = result.get('behavioral', {})
    rows_v2 = []
    for a in b.get('actions', []):
        rows_v2.append([a.get('id',''), a.get('name',''), a.get('input',''), a.get('output',''), a.get('domain','')])
    for sm in b.get('stateMachines', []):
        for t in sm.get('transitions', []):
            rows_v2.append([sm.get('id',''), sm.get('name',''), t.get('from',''), t.get('to',''), t.get('trigger','')])
    for ind in b.get('indicators', []):
        rows_v2.append([ind.get('id',''), ind.get('name',''), ind.get('formula',''), ind.get('target',''), ind.get('warningThreshold','')])
    ws3 = wb.create_sheet('维2_动态行为')
    _write_sheet(ws3, ['ID', '名称', '输入/From', '输出/To', '触发/目标'], rows_v2)

    # Sheet 4: 维3_规则约束
    r = result.get('rules', {})
    rows_v3 = []
    for v in r.get('validations', []):
        rows_v3.append(['校验', v.get('id',''), v.get('type',''), v.get('entity',''), v.get('field',''), v.get('expression','')])
    for g in r.get('guardrails', []):
        rows_v3.append(['护栏', g.get('id',''), g.get('name',''), g.get('condition',''), g.get('action',''), ''])
    for p in r.get('policies', []):
        rows_v3.append(['策略', p.get('id',''), p.get('name',''), p.get('rules',''), '', ''])
    for perm in r.get('permissions', []):
        rows_v3.append(['权限', perm.get('role',''), perm.get('resource',''), perm.get('operations',''), '', ''])
    for pr in r.get('probes', []):
        rows_v3.append(['探针', pr.get('id',''), pr.get('name',''), pr.get('target',''), pr.get('frequency',''), pr.get('alertCondition','')])
    ws4 = wb.create_sheet('维3_规则约束')
    _write_sheet(ws4, ['类型', 'ID', '名称/表达式', '作用域', '字段/条件', '补充'], rows_v3)

    # Sheet 5: 维4_事件消息
    e = result.get('events', {})
    rows_v4 = []
    for et in e.get('eventTypes', []):
        rows_v4.append([et.get('id',''), et.get('name',''), et.get('severity',''), et.get('source',''), et.get('targetEntity','')])
    for ca in e.get('causalities', []):
        rows_v4.append(['因果链', ca.get('cause',''), ca.get('effect',''), '', ''])
    ws5 = wb.create_sheet('维4_事件消息')
    _write_sheet(ws5, ['事件ID', '名称', '级别', '事件源', '作用实体'], rows_v4)

    # Sheet 6: 维5_外部接口
    i = result.get('interfaces', {})
    rows_v5 = []
    for api in i.get('apis', []):
        rows_v5.append(['API', api.get('id',''), api.get('name',''), api.get('url',''), api.get('method','')])
    for q in i.get('queries', []):
        rows_v5.append(['查询', q.get('id',''), q.get('name',''), q.get('type',''), q.get('template','')])
    for comp in i.get('compute', []):
        rows_v5.append(['计算', comp.get('id',''), comp.get('name',''), comp.get('input',''), comp.get('formula','')])
    for n in i.get('notifications', []):
        rows_v5.append(['通知', n.get('id',''), n.get('name',''), n.get('channel',''), n.get('template','')])
    for rep in i.get('reports', []):
        rows_v5.append(['报表', rep.get('id',''), rep.get('name',''), rep.get('format',''), rep.get('fields','')])
    ws6 = wb.create_sheet('维5_外部接口')
    _write_sheet(ws6, ['类型', 'ID', '名称', 'URL/渠道/公式', '方法/模板'], rows_v5)

    # Sheet 7: EPC流程
    epc = result.get('epc', {})
    rows_epc = []
    for step in epc.get('steps', []):
        rows_epc.append([step.get('event_trigger',''), step.get('action',''), ','.join(step.get('conditions', [])), ','.join(step.get('guards', []))])
    ws7 = wb.create_sheet('EPC流程')
    _write_sheet(ws7, ['触发事件', '行为', '条件', '护栏'], rows_epc)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@bp_export.get('/xlsx/<release_no>')
@require_auth()
def export_xlsx(release_no):
    result, rel = _get_release_models(release_no)
    if result is None:
        return jsonify({'error': 'not found'}), 404
    output = _build_xlsx(result)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp_export.route('/xlsx-from-manifest', methods=['GET', 'POST'])
@require_auth()
def export_xlsx_from_manifest():
    """直接从 manifest JSON 生成 Excel（无需 release_no）。"""
    payload = request.get_json(force=True)
    result = {
        'version': payload.get('version', payload.get('metadata', {}).get('version', '0.0.0')),
        'exported_at': None,
        'structural': payload.get('structural', payload.get('data', {})),
        'behavioral': payload.get('behavioral', {}),
        'rules': payload.get('rules', {}),
        'events': payload.get('events', {}),
        'interfaces': payload.get('interfaces', {}),
        'epc': payload.get('epc', {}),
    }
    output = _build_xlsx(result)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
